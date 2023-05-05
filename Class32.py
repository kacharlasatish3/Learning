import openpyxl
import pytest
import json
import random
import string
import requests
from configurations import *

# @pytest.mark.usefixtures("get_authorization_token")
def create_user():
    token = "11232"
    url = baseurl
    header = {"Content-Type": "application/json",
              "Authorization": token}
    #
    # name = ''.join(random.choices(string.ascii_lowercase + string.ascii_uppercase +
    #                               string.digits, k=10))
    # email = name + "@gmail.com"
    # data = json.dumps({
    #     "name": name,
    #     "email": email,
    #     "gender": "male",
    #     "status": "active"
    # })
    # response = requests.post(url, headers=header, data=data)
    # assert response.status_code == 201
    # result = response.json()
    # assert result["id"] is not None
    # assert result["name"] == name
    # assert result["email"] == email
    # assert result["gender"] == "male"
    # assert result["status"] == "active"
    path = "Book1.xlsx"

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    # max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    # Loop will print all columns name
    for i in range(1,max_row+1):

        name = sheet_obj.cell(row=i, column=1)
        email =sheet_obj.cell(row=i, column=2)
        gender = sheet_obj.cell(row=i,column=3)
        # for j in range(1, max_col + 1):
        #     cell_obj = sheet_obj.cell(row=i, column=j)
        #     print(cell_obj.value,end=" ")
        # print("")
        data = json.dumps({
            "name": name,
            "email": email,
            "gender": gender,
            "status": "active"
        })
        response = requests.post(url, headers=header, data=data)
        assert response.status_code == 201
        result = response.json()
        assert result["id"] is not None
        assert result["name"] == name
        assert result["email"] == email
        assert result["gender"] == "male"
        assert result["status"] == "active"


create_user()